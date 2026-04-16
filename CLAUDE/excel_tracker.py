from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = ["Date", "Medicine", "Dosage", "Scheduled Time", "Taken At", "Status"]
HEADER_BG = "4F81BD"
COLUMN_WIDTHS = [12, 22, 15, 16, 22, 10]


def _create_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dose Tracker"

    for col, (header, width) in enumerate(zip(HEADERS, COLUMN_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(path)


def log_dose_to_excel(excel_path: Path, medicine: dict) -> None:
    """Append a single dose record to the Excel tracker."""
    if not excel_path.exists():
        _create_workbook(excel_path)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    now = datetime.now()
    ws.append([
        now.strftime("%Y-%m-%d"),
        medicine.get("name", ""),
        medicine.get("dosage", ""),
        medicine.get("time", ""),
        now.strftime("%Y-%m-%d %H:%M:%S"),
        "Taken",
    ])

    wb.save(excel_path)
